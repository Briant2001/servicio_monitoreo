"""
Helper to save indicators (press, temperature_c, humidity_rh) to CSV and optional XLSX.
Usage from your `app.py`:

    from save_indicaciones import save_indicacion
    save_indicacion(press, temperature_c, humidity_rh, csv_path='indicaciones.csv', xlsx_path='indicaciones.xlsx')

If you want XLSX support, install `openpyxl` in your environment: `pip install openpyxl`.
"""

import csv
import os
from datetime import datetime
import sys
from resource_path import resource_path
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except Exception:
    openpyxl = None


def append_to_csv(file_path, press_cruda, press, temperature_c, humidity_rh, temperature_c_2, timestamp=None):
    """Append a row to a CSV file, creating it with headers if missing."""
    if timestamp is None:
        timestamp = datetime.now().isoformat()

    dirpath = os.path.dirname(file_path)
    if dirpath and not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)

    file_exists = os.path.exists(file_path)
    header = [
        "timestamp",
        "press_cruda",
        "press",
        "temperature_c",
        "humidity_rh",
        "temperature_c_2",
        "humidity_rh_cruda"
    ]

    with open(file_path, mode="a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(header)
        writer.writerow([timestamp, press_cruda, press, temperature_c, humidity_rh, temperature_c_2,])


def append_to_xlsx(file_path, press, temperature_c, humidity_rh,temperature_c_2, timestamp=None):
    """Append a row to an XLSX file (requires openpyxl). Creates workbook if missing."""
    if openpyxl is None:
        raise ImportError("openpyxl is required to write XLSX files. Install with: pip install openpyxl")

    if timestamp is None:
        timestamp = datetime.now().isoformat()

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["timestamp", "press", "temperature_c", "temperature_c_2"])  # header

    ws.append([timestamp, press, temperature_c, temperature_c_2])
    wb.save(file_path)


def save_indicacion(press_cruda, press, temperature_c, humidity_rh, temperature_c_2, csv_path="Condiciones Ambientales.csv", xlsx_path=None, timestamp=None):
    """
    Save the provided indicators to CSV and optionally to XLSX.

    Returns a dict with status information.
    """
    result = {"csv": False, "xlsx": False, "errors": []}

    try:
        append_to_csv(csv_path, press_cruda, press, temperature_c, humidity_rh, temperature_c_2, timestamp=timestamp)
        result["csv"] = True
    except Exception as e:
        result["errors"].append(f"CSV error: {type(e).__name__} - {e}")

    if xlsx_path:
        try:
            append_to_xlsx(xlsx_path, press, temperature_c, humidity_rh, timestamp=timestamp)
            result["xlsx"] = True
        except Exception as e:
            result["errors"].append(f"XLSX error: {type(e).__name__} - {e}")

    return result
